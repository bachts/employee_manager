import pandas as pd
import os
import openpyxl
from django.db import connection
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import range_boundaries

def GenerateExcelSheet(basedir, levels,data_dictionary) -> None:
        cursor = connection.cursor()
        # print("dictionary_data: ",data_dictionary)
        sql_query='''select ee.id as employeeId,
                            full_name as name,
                            level,
                            ee.team_id as teamId,
                            et.name as teamName,
                            type,
                            oo.kr_id as krId,
                            oo.key_result_department as krDep,
                            oo.key_result_team as krTeam,
                            oo.key_result_personal as krPer,
                            ofo.formula_name as formulaName,
                            ofo.formula_value as formulaValue,
                            os.source_name as sourceName,
                            oo.regularity,
                            oo.unit,
                            oo.condition,
                            oo.norm,
                            oo.weight,
                            oo.result,
                            oo.weight*oo.result as ratio,
                            oo.estimated,
                            oo.actual,
                            oo.note
                            FROM "Employee_employee" as ee,
                                "Employee_team" as et,
                                "OKR_okr" as oo,
                                "OKR_formula" as ofo,
                                "OKR_source" as os,
                                "Employee_department" as ed
                            where ee.team_id=et.team_id
                                and ee.id=oo.user_id
                                and oo.formula_id=ofo.id
                                and oo.source_id=os.id
                                
                    '''
        # Kiểm tra và thêm điều kiện từ dictionary
        if data_dictionary.get("quater")!='':
            sql_query += " and oo.deadline_quarter = %(quater)s"
        if data_dictionary.get("month")!='':
            sql_query += " and oo.deadline_month = %(month)s"
        if data_dictionary.get("year")!='':
            sql_query += " and oo.deadline_year = %(year)s"
        if data_dictionary.get("department_id")!='':
            sql_query += " and ed.department_id = %(department_id)s"
   
        cursor.execute(sql_query, data_dictionary)
        result = cursor.fetchall()
        dataframe = pd.DataFrame(result)
        # print("các column: ", dataframe)

        dataframe.columns = ['employeeId', 'Name', 'level', 'teamId', 'teamName',
                        'Loại', 'krId', 'KR phòng', 'KR team', 'KR cá nhân', 'Công thức tính', 'Giá trị tính',
                        'Nguồn dữ liệu', 'Định kỳ tính', 'Đơn vị tính', 'Điều kiện', 'Norm',
                        '% Trọng số chỉ tiêu', 'Kết quả', 'Tỷ lệ', 'Tổng thời gian dự kiến/ ước tính công việc (giờ)',
                            'Tổng thời gian thực hiện công việc thực tế (giờ)', 'Note']

        dataframe.sort_values('employeeId', inplace=True)
        # dataframe.drop(columns=[ei], axis=1, inplace=True)
        dataframe.drop(columns=['teamId'], axis=1, inplace=True)
        dataframe.drop(columns=['Giá trị tính'], axis=1, inplace=True)
        dataframe.replace("NUM", "%", inplace=True)
        dataframe.replace("CAT", "Đạt/Không đạt", inplace=True)
        dataframe.replace("MO", "Tháng", inplace=True)
        dataframe.replace("QUAR", "Quý", inplace=True)
        dataframe.replace("EQUAL", "=", inplace=True)
        dataframe.fillna(0, inplace=True)
        # dataframe.set_index(['id', 'full_name', 'department_name', 'team_name', 'type', 'okr_kpi_id', 'objective_name', 'type'], inplace=True)
        for value, str in levels.items():
            temp_df = dataframe.loc[dataframe['level'] == value]
            temp_df.drop(columns=['level'], axis=1)
            file_directory = basedir+f"\\nhóm {str}.xlsx"
            if os.path.exists(file_directory):
                os.remove(file_directory)
            if not temp_df.empty:
                temp_df.to_excel(file_directory)
                formatKPIExcelSheet(file_directory)
                


def excelToDataframe(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active
    data = sheet.values
    columns = next(data)  # Lấy tên cột từ hàng đầu tiên
    # Tạo DataFrame từ dữ liệu
    df = pd.DataFrame(data, columns=columns)
    wb.close()
    return df


def formatKPIExcelSheet(file_path) -> None:

    # Tạo DataFrame từ dữ liệu
    df = excelToDataframe(file_path)
    # print("dataframe sorted_df: \n",df)
    # tạo biến tên cho các trường trong dataframe để sử dụng sau
    krp = 'KR phòng'
    krt = 'KR team'
    krc = 'KR cá nhân'
    ctt = 'Công thức tính'
    ei = 'employeeId'
    ki = 'krId'
    type = 'Loại'
    name = 'Name'
    lv = 'level'
    tn = 'teamName'
    note = 'Note'
    tsct = '% Trọng số chỉ tiêu'
    kq = 'Kết quả'
    tl = 'Tỷ lệ'
    et = 'Tổng thời gian dự kiến/ ước tính công việc (giờ)'
    rt = 'Tổng thời gian thực hiện công việc thực tế (giờ)'
    df[tsct] = df[tsct].astype(int)
    df[kq] = df[kq].astype(float)
    df[tl] = df[tl].astype(float)
    df[et] = df[et].astype(float)
    df[rt] = df[rt].astype(float)
    # tính tổng các trường cần tính
    tsct_sum = df.groupby([ei, ki])[tsct].sum().reset_index()
    kq_sum = df.groupby([ei, ki])[kq].sum().reset_index()
    tl_sum = df.groupby([ei, ki])[tl].sum().reset_index()
    et_sum = df.groupby([ei])[et].sum().reset_index()
    rt_sum = df.groupby([ei])[rt].sum().reset_index()
    # Tạo DataFrame chứa các tổng
    df_data_sum = pd.DataFrame({ei: tsct_sum[ei],
                           ki: tsct_sum[ki],
                           tsct: tsct_sum[tsct],
                           kq: kq_sum[kq],
                           tl: tl_sum[tl]})
    df_time_sum = pd.DataFrame({ei: et_sum[ei],
                           et: et_sum[et],
                           rt: rt_sum[rt]})

    # merge các tổng đã tính ở trên vào một hàng và tìm vị trị cần điền các tổng đó trong dataframe gốc
    max_indices = df_data_sum.groupby(ei)[ki].idxmax()
    merged_sum_idx_df = pd.merge(
        max_indices, df_time_sum, on=ei, how='left')
    merged_sum_idx_df.drop(columns=[ei], axis=1, inplace=True)
    krId_to_idx = merged_sum_idx_df.set_index(ki)
    merged_sum_df = pd.merge(df_data_sum, krId_to_idx,
                             left_index=True, right_index=True, how='left')
    merged_sum_df.fillna(0, inplace=True)
    merged_sum_df[et] = merged_sum_df[et].apply(lambda x: '' if x == 0 else x)
    merged_sum_df[rt] = merged_sum_df[rt].apply(lambda x: '' if x == 0 else x)
    # print("đây là merged_sum_df: \n",merged_sum_df)

    # sắp xếp lại và lấy vị trí các đoạn cần insert các tổng vào (need_add_index_df), đồng thời dùng sorted_df cho các thao tác sau (sorted_df)
    sorted_df = df.sort_values(by=[ei, ki], ascending=[
                               True, True]).reset_index()
    sorted_df.drop(columns=['index'], axis=1, inplace=True)
    # print("dataframe sorted_df: \n",sorted_df)
    index_sorted_df = sorted_df[[ei, ki]]
    # Tạo một cột boolean cho biết các hàng có là bản sao của hàng trước đó hay không
    is_duplicated = index_sorted_df.duplicated(subset=[ei, ki], keep='last')
    # Chỉ giữ lại các hàng cuối cùng của mỗi cặp giá trị bằng nhau
    need_add_index_df = index_sorted_df[~is_duplicated]
    # print("đây là need_add_index_df : \n",need_add_index_df)
    # thêm các kết quả đã tính toán ở trên vào cuối của mỗi kpi
    added_row = 0
    map_new_index = {}
    for (index, i) in zip(need_add_index_df.index, range(len(merged_sum_df))):
            df_new_record = pd.DataFrame(merged_sum_df.iloc[i, :]).T
            true_position = index+added_row+1
            sorted_df = pd.concat([sorted_df.iloc[:true_position], df_new_record,
                                  sorted_df.iloc[true_position:]]).reset_index(drop=True)
            # thêm các giá trị index mới vào dictionary
            map_new_index[index] = true_position
            added_row += 1
    sorted_df.drop(columns=[None], axis=1, inplace=True)
    sorted_df = sorted_df.apply(lambda x: '' if x.empty else x)
    need_add_index_df.set_index(need_add_index_df.index.map(map_new_index), inplace=True)
    # print("đây là need_add_index_df mới : \n",need_add_index_df)
    # print("đây là sorted_df mới : \n",sorted_df)

    # 1. tạo ra các dataframe chứa tên, level, tên nhóm và dataframe chưa tên các cột  (user_df)
    grouped = sorted_df.groupby([ei, name, lv, tn])
    user_df = grouped.size().reset_index(name='Count')
    user_df.insert(0, 'Index', user_df.index+1)
    user_df.drop(columns=['Count'], axis=1, inplace=True)
    # print("đây là user_df : \n",user_df)

    # Create a new DataFrame to store the column names (title_df)'
    columns_to_drop = [ei, name, lv, tn, ki]
    insert_header_name_df = sorted_df.drop(columns=columns_to_drop, axis=1)
    title_df = pd.DataFrame([insert_header_name_df.columns],
                            columns=insert_header_name_df.columns)
    # print("đây là user_df : \n",title_df)

    # 2. tạo ra một list chứa các vị trí cần add các record đó vào bằng cách tìm các vị trí đầu xuất hiện của các record
    # Tạo cột mới để xác định lần xuất hiện đầu tiên của giá trị trùng nhau
    sorted_df['FirstOccurrence'] = ~sorted_df.duplicated(subset=[ei])
    # lấy lần xuất hiện đầu tiên của các giá trị trùng nhau (firstOccurrence_df)
    firstOccurrence_df = sorted_df[sorted_df['FirstOccurrence']].drop(
        columns='FirstOccurrence')
    firstOccurrence_list = firstOccurrence_df.index
    sorted_df = sorted_df.drop(columns='FirstOccurrence', axis=1)
    
    # 3. dùng openpyxl thêm các header vào các vị trí(2.) là:
    #                                      - các record(1.)
    #                                      - các record tên cột(1.)
    #                                      - Tên nhóm là header chính

    # Chuyển DataFrame thành một đối tượng Sheet bằng pandas dataframe_to_rows
    user_rows = list(dataframe_to_rows(user_df, index=False, header=False))
    if len(user_rows) != len(firstOccurrence_list):
        raise ValueError("Số lượng dòng cần thêm phải bằng số dự liệu muốn thêm!")
    
    # tạo ra 1 dataframe có 3 hàng là tên các cột cần add vào sheet
    for i in range(len(user_rows)):
         title_df = pd.concat([title_df, title_df], ignore_index=True)

    # lấy vị trí các cột cần mở rộng khi align bằng openpyxl
    sum_column_names = [tsct, kq, tl, et,rt]
    sum_column_positions = [title_df.columns.get_loc(col_name) for col_name in sum_column_names]

    column_name_rows = list(dataframe_to_rows(title_df, index=False, header=False))
    rows = dataframe_to_rows(insert_header_name_df, index=False, header=False)

# ---------------------------------------------------------------------------------
    # tạo một sheet excel mới
    workbook = Workbook()
    sheet = workbook.active
    # Ghi dữ liệu từ DataFrame vào Workbook
    try:
        level = sorted_df[lv][0].astype(int).astype(str)
    except ValueError:
        print("Level đang không là số!")
        level = sorted_df[lv][0].astype(str)
    # Gán tiêu đề cho sheet
    Header_font = Font(name='Arial',size=16, bold=True)
    if(level==0):
        header_text="nhóm SVCNTS"
    else:
        header_text=f"nhóm L{level}"
    sheet.cell(row=1, column=1, value=header_text).font = Header_font
    # Merge các ô trong dòng đầu tiên
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=title_df.shape[1])

    # Chèn các dòng vào vị trí xác định (từ dòng 2 trở đi)
    data_font = Font(name='Arial',size=11, bold=False)
    for r_idx, row in enumerate(rows, 2):
        for c_idx, value in enumerate(row, 1):
            sheet.cell(row=r_idx, column=c_idx, value=value)
            sheet.cell(row=r_idx, column=c_idx, value=value).font = data_font

    # thêm các dòng thông tin người dung và title của các trường vào sheet, đồng thời thêm màu, sửa font
    light_blue_fill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
    dark_blue_fill = PatternFill(start_color='365072', end_color='365072', fill_type='solid')
    light_yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
    light_green_fill = PatternFill(start_color='A4C639', end_color='A4C639', fill_type='solid')
    dark_green_fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
    user_title_font = Font(name='Arial',size=11, bold=True,color="FF0000")
    title_font = Font(name='Arial',size=11, bold=True,color="FFFFFF")
    
    # lấy vị trí các column cần đổi màu xanh lá cây nhạt
    column_light_green_fill = [tsct, kq, tl, et, note]
    column_positions_light_green_fill = [title_df.columns.get_loc(col_name) for col_name in column_light_green_fill]

    # # lấy vị trí các column cần đổi màu xanh lá cây đậm
    # column_dark_green_fill = [tsct, 'KR team', 'KR cá nhân', 'Công thức tính', note]
    # column_positions_dark_green_fill = [title_df.columns.get_loc(col_name) for col_name in column_dark_green_fill]

    # lấy vị trí các column cần đổi màu xanh da trời nhạt
    column_light_blue_fill = [rt]
    column_positions_light_blue_fill = [title_df.columns.get_loc(col_name) for col_name in column_light_blue_fill]

    for row_sum_index in need_add_index_df.index:
        for col_sum_index in sum_column_positions:
            # print("col and row",col_sum_index,row_sum_index)
            # ở đây row phải +2 là vì: sheet có các index từ số 1 + header là lấy mất 1 dòng đầu
            # ở đây col phải +1 là vì: sheet có các index từ số 1
            sheet.cell(row=row_sum_index+2, column=col_sum_index+1).fill = light_yellow_fill

    added_sheet_row = 0
    for insert_index, row_user_data,row_column_name in zip(firstOccurrence_list, user_rows,column_name_rows):       
        user_insert = insert_index+added_sheet_row+2
        title_insert = insert_index+added_sheet_row+3
        sheet.insert_rows(user_insert, 1)
        sheet.insert_rows(title_insert, 1)
        # thêm dòng thông tin user
        for col_idx_user,user_value in enumerate(row_user_data, 1):
            sheet.cell(row=user_insert, column=col_idx_user, value=user_value)
            sheet.cell(row=user_insert, column=col_idx_user).fill = light_blue_fill
            sheet.cell(row=user_insert, column=col_idx_user).font = user_title_font
        # thêm dòng title
        for col_idx,column_name_value in enumerate(row_column_name,1):
            sheet.cell(row=title_insert, column=col_idx, value=column_name_value)
            sheet.cell(row=title_insert, column=col_idx).fill = dark_blue_fill
            sheet.cell(row=title_insert, column=col_idx).font = title_font
            for green_index in column_positions_light_green_fill:
                if green_index==col_idx:
                    sheet.cell(row=title_insert, column=green_index).fill = light_green_fill
            for blue_index in column_positions_light_blue_fill:
                if blue_index==col_idx:
                    sheet.cell(row=title_insert, column=blue_index).fill = light_blue_fill

        if(insert_index!=0):
             blank_insert = insert_index+added_sheet_row+2
             sheet.insert_rows(blank_insert, 1) 
             added_sheet_row+=3       
        else:
             added_sheet_row+=2

    # Biến trạng thái
    merging = False
    merge_start = None

    # lấy vị trí cột loại để merge các cell 
    merge_column_position = title_df.columns.get_loc(type)

    # Duyệt qua các dòng trong cột "loại"
    for row_index, row in enumerate(sheet.iter_rows(min_col=merge_column_position, max_col=merge_column_position, values_only=True), start=1):
        if row[0] =='KPI' or row[0]=='OKR':
            if not merging:
                merging = True
                merge_start = row_index
        else : 
            if merging:
                merging = False
                sheet.merge_cells(start_row=merge_start, start_column=1, end_row=row_index - 1, end_column=1)

    # Thiết lập tự động tăng kích thước các cột
    # for col in sheet.columns:
    #     max_length = 0
    #     for cell in col:
    #         try:
    #             if len(str(cell.value)) > max_length:
    #                 max_length = len(cell.value)
    #         except:
    #             pass
    #     adjusted_width = (max_length + 2) * 1.2
    #     sheet.column_dimensions[col[0].column_letter].width = adjusted_width

    # Đặt viền cho các cell
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))
    # Loop through each worksheet in the workbook
    for sheet in workbook.worksheets:
        # Loop through each row in the worksheet
        for row in sheet.iter_rows():
            # Loop through each cell in the row
            for cell in row:
                # Set wrap_text to True to automatically wrap text
                cell.alignment = openpyxl.styles.Alignment(wrapText=True)
                # thêm viền
                cell.border = border
                # Thiết lập tự động align vào giữa cho toàn bộ text trong sheet
                cell.alignment = Alignment(wrap_text=True,horizontal='center', vertical='center')

        # Auto-adjust row height for all rows in the worksheet
        for row in sheet.iter_rows():
            sheet.row_dimensions[row[0].row].auto_size = True
        
    # Define a dictionary to store the desired column widths
    # column_widths = {'KR phòng': 20, 'KR team': 20, 'KR cá nhân': 20, 'Note': 20}  

    # lấy vị trí các cột cần mở rộng khi align bằng openpyxl
    column_names_30 = [krp, krt, krc, ctt, note]
    column_positions_30 = [title_df.columns.get_loc(col_name) for col_name in column_names_30]

    column_names_20 = [ type, et, rt]
    column_positions_20 = [title_df.columns.get_loc(col_name) for col_name in column_names_20]

    # Update the column widths
    for index in column_positions_30:
        # đoạn này index +1 là vì vào sheet excel các index tính từ 1 chứ không còn là 0 như ở dataframe
        column_letter = openpyxl.utils.get_column_letter(index+1)
        sheet.column_dimensions[column_letter].width = 30
    # Update the column widths
    for index in column_positions_20:
        # đoạn này index +1 là vì vào sheet excel các index tính từ 1 chứ không còn là 0 như ở dataframe
        column_letter = openpyxl.utils.get_column_letter(index+1)
        sheet.column_dimensions[column_letter].width = 20
        
    # Đọc nội dung của sheet thành DataFrame
    # data = sheet.values
    # columns = next(data)  # Lấy tên cột từ dòng đầu tiên
    # final_df = pd.DataFrame(data, columns=columns)  
    # print("đây là excel file cuối : \n",final_df)

    # Lưu Workbook
    workbook.close()
    workbook.save(file_path)

def synthesizeExcelFilebySheet(listDirectory,targetDirectory,levels) -> None:

    # Tạo một workbook mới để tổng hợp dữ liệu
    wb_combined = openpyxl.Workbook()
    # Lặp qua từng file Excel gốc
    for file_name,level in zip(listDirectory,levels):
        # Mở file Excel gốc
        wb_original = openpyxl.load_workbook(file_name)
        # Chọn sheet trong file Excel gốc 
        sheet_original = wb_original.active 
        for level in levels:
            # Tạo một sheet mới trong workbook tổng hợp có title phù hợp với level của nhân viên trong sheet
            title=''
            if(level==0):
                title="nhóm SVCNTS"
            else:
                title=f"nhóm L{level}"
            if title in file_name:   
                sheet_combined = wb_combined.create_sheet(title)
            else:
                continue        
        # Lặp qua từng hàng và cột trong sheet gốc
        for row in sheet_original.iter_rows():
            for cell in row:
                # Copy giá trị sang sheet tổng hợp
                new_cell = sheet_combined.cell(row=cell.row, column=cell.column, value=cell.value)
                
                # Sao chép các thuộc tính định dạng
                new_cell.font = openpyxl.styles.Font(
                    name=cell.font.name,
                    size=cell.font.size,
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    color=cell.font.color,
                    underline=cell.font.underline
                )
                
                new_cell.fill = openpyxl.styles.PatternFill(
                    fill_type=cell.fill.fill_type,
                    start_color=cell.fill.start_color,
                    end_color=cell.fill.end_color
                )
                
                new_cell.border = openpyxl.styles.Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom
                )
                
                new_cell.alignment = openpyxl.styles.Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical=cell.alignment.vertical,
                    text_rotation=cell.alignment.text_rotation,
                    wrap_text=cell.alignment.wrap_text,
                    shrink_to_fit=cell.alignment.shrink_to_fit,
                    indent=cell.alignment.indent
                )

        # Sao chép độ rộng cột từ sheet gốc sang sheet tổng hợp
        for col in sheet_original.column_dimensions:
            sheet_combined.column_dimensions[col] = sheet_original.column_dimensions[col]

        # Lấy danh sách các vùng merge
        merged_ranges = sheet_original.merged_cells.ranges
        # print("merged_ranges:",merged_ranges)
        # Loop through each merged cell range
        for merged_range in merged_ranges:
            # Extract the start and end row indices from the range string
            start_col, start_row, end_col, end_row = range_boundaries(merged_range.coord)
            # print("coordination:",start_row, start_col, end_row, end_col)
            
            # Merge the cells in the range
            sheet_combined.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)

        # # Duyệt qua các cell trong worksheet cũ
        # for row in sheet_original.iter_rows(min_row=1, max_row=sheet_original.max_row, min_col=1, max_col=sheet_original.max_column):
        #     for cell in row:
        #         # Nếu cell thuộc vùng merge
        #         if cell.coordinate in merged_ranges:
        #             # Lấy giá trị của vùng merge
        #             merged_value = sheet_original[cell.coordinate].value
        #             # Gán giá trị của vùng merge cho tất cả các cell trong vùng merge
        #             for merged_cell in sheet_original[merged_ranges[0]]:
        #                 sheet_combined.cell(row=merged_cell.row, column=merged_cell.column, value=merged_value)

    # Xóa sheet mặc định trong Workbook tổng hợp
    if 'Sheet' in wb_combined.sheetnames:
        wb_combined.remove(wb_combined['Sheet'])
    # Lưu workbook tổng hợp lại vào một file mới
    wb_combined.save(targetDirectory)


# formatKPIExcelSheet("excelSheet/nhóm L2.xlsx")


# def department() -> None:
#     column_order = ['okr_kpi_id', 'full_name', 'department_name', 'type',
#                     'objective_name', 'kr_phong', 'kr_team', 'kr_personal', 'unit',
#                     'condition', 'norm', 'weight', 'ratio',  'result', 'status', 'deadline'] 

#     department_df = full_df[column_order]
#     department_df['deadline'] = department_df['deadline'].dt.tz_localize(None)

#     def do_nothing(x):
#         return x

#     # department_df.set_index(['department_name', 'type'], inplace=True)
#     # print(department_df)
#     department_df = department_df.fillna('No data')
#     # department_df = department_df.groupby(['department_name', 'type', 'okr_kpi_id', 'objective_name', 'weight']).agg({
#     #     'full_name': do_nothing, # lambda x: ', '.join(x),
#     #     'kr_phong': do_nothing,
#     #     'kr_team': do_nothing,
#     #     'unit': do_nothing,
#     #     'condition': do_nothing,
#     #     'norm': do_nothing,
#     #     'weight': do_nothing,
#     #     'ratio': do_nothing,
#     #     'result': do_nothing,
#     #     'deadline': do_nothing,
#     # })
#     department_df = department_df.set_index(['department_name', 'type', 'okr_kpi_id', 'objective_name', 'kr_phong', 'kr_team'])
        # dataframe.rename(columns={"0": "employeeId"}, inplace=True)
        # dataframe.rename(columns={"1": "fullName"}, inplace=True)
        # dataframe.rename(columns={"2": "level"}, inplace=True)
        # dataframe.rename(columns={"3": "teamId"}, inplace=True)
        # dataframe.rename(columns={"4": "teamName"}, inplace=True)
        # dataframe.rename(columns={"5": "Loại"}, inplace=True)
        # dataframe.rename(columns={"6": "KR phòng"}, inplace=True)
        # dataframe.rename(columns={"7": "KR team"}, inplace=True)
        # dataframe.rename(columns={"8": "KR cá nhân"}, inplace=True)
        # dataframe.rename(columns={"9": "Công thức tính"}, inplace=True)
        # dataframe.rename(columns={"10": "Nguồn dữ liệu"}, inplace=True)
        # dataframe.rename(columns={"11": "Định kỳ tính"}, inplace=True)
        # dataframe.rename(columns={"12": "Đơn vị tính"}, inplace=True)
        # dataframe.rename(columns={"13": "Điều kiện"}, inplace=True)
        # dataframe.rename(columns={"14": "Norm"}, inplace=True)
        # dataframe.rename(columns={"15": "% Trọng số chỉ tiêu"}, inplace=True)
        # dataframe.rename(columns={"16": "Kết quả"}, inplace=True)
        # dataframe.rename(columns={"17": "Tỷ lệ"}, inplace=True)
        # dataframe.rename(columns={"18": "Tổng thời gian dự kiến/ ước tính công việc (giờ)"}, inplace=True)
        # dataframe.rename(columns={"19": "Tổng thời gian thực hiện công việc thực tế (giờ)"}, inplace=True)
        # dataframe.rename(columns={"20": "Note"}, inplace=True)

#     department_df.to_excel('department.xlsx')

# def okr_quarter() -> None:

    

#     quarter_df = full_df.loc[full_df['type'] == 'okr']
#     quarter_df['deadline'] = quarter_df['deadline'].dt.tz_localize(None)
#     column_order = ['okr_kpi_id', 'id', 'full_name', 'department_name', 'team_name',
#                     'objective_name', 'kr_phong', 'regularity', 'unit', 'condition',
#                     'result', 'deadline', 'status', 'files']
#     quarter_df = quarter_df[column_order]
#     quarter_df = quarter_df.fillna('No data')
#     quarter_df = quarter_df.set_index(['department_name', 'team_name', 'id', 'full_name', 'okr_kpi_id', 'objective_name'])
#     quarter_df.to_excel('quarter.xlsx')


